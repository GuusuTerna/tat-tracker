from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///tat.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your-secret-key-here'

db = SQLAlchemy(app)

# ========== MODELS (in correct order - SLAConfig FIRST) ==========

class SLAConfig(db.Model):
    __tablename__ = 'sla_configs'
    
    id = db.Column(db.Integer, primary_key=True)
    test_name = db.Column(db.String(100), unique=True)
    department = db.Column(db.String(50))
    sla_hours = db.Column(db.Float)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class Case(db.Model):
    __tablename__ = 'cases'
    
    id = db.Column(db.Integer, primary_key=True)
    visit_id = db.Column(db.String(50))
    case_id = db.Column(db.String(20), unique=True)
    patient_name = db.Column(db.String(100))
    department = db.Column(db.String(50))
    test_name = db.Column(db.String(100))
    test_category = db.Column(db.String(50))
    status = db.Column(db.String(50), default='Not Started')
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    result_ready_at = db.Column(db.DateTime)
    
    # Lab phases
    sample_collection_start = db.Column(db.DateTime)
    sample_collection_end = db.Column(db.DateTime)
    sample_labeling_start = db.Column(db.DateTime)
    sample_labeling_end = db.Column(db.DateTime)
    sample_transport_start = db.Column(db.DateTime)
    sample_transport_end = db.Column(db.DateTime)
    sample_reception_start = db.Column(db.DateTime)
    sample_reception_end = db.Column(db.DateTime)
    sample_processing_start = db.Column(db.DateTime)
    sample_processing_end = db.Column(db.DateTime)
    analysis_start = db.Column(db.DateTime)
    analysis_end = db.Column(db.DateTime)
    review_start = db.Column(db.DateTime)
    review_end = db.Column(db.DateTime)
    validation_start = db.Column(db.DateTime)
    validation_end = db.Column(db.DateTime)
    result_upload_start = db.Column(db.DateTime)
    result_upload_end = db.Column(db.DateTime)
    
    # Radiology phases
    exam_start = db.Column(db.DateTime)
    exam_end = db.Column(db.DateTime)
    post_exam_start = db.Column(db.DateTime)
    post_exam_end = db.Column(db.DateTime)
    
    # Mobile phases
    register_start = db.Column(db.DateTime)
    register_end = db.Column(db.DateTime)
    mobile_collect_start = db.Column(db.DateTime)
    mobile_collect_end = db.Column(db.DateTime)
    transport_to_lab_start = db.Column(db.DateTime)
    transport_to_lab_end = db.Column(db.DateTime)
    deliver_results_start = db.Column(db.DateTime)
    deliver_results_end = db.Column(db.DateTime)
    
    # Histology
    fixation_done = db.Column(db.Boolean, default=False)
    fixation_date = db.Column(db.DateTime)
    embedding_done = db.Column(db.Boolean, default=False)
    embedding_date = db.Column(db.DateTime)
    cutting_done = db.Column(db.Boolean, default=False)
    cutting_date = db.Column(db.DateTime)
    staining_done = db.Column(db.Boolean, default=False)
    staining_date = db.Column(db.DateTime)
    histology_notes = db.Column(db.Text)
    
    manual_durations = db.Column(db.Text, default='{}')
    tat_hours = db.Column(db.Float)
    sla_status = db.Column(db.String(20), default='PENDING')
    
    def get_manual_durations_dict(self):
        if self.manual_durations:
            return json.loads(self.manual_durations)
        return {}
    
    def set_manual_duration(self, phase, minutes):
        durations = self.get_manual_durations_dict()
        durations[phase] = minutes
        self.manual_durations = json.dumps(durations)
        db.session.commit()
    
    def get_sla_hours(self):
        """Get SLA hours for this test - checks custom config first"""
        try:
            custom = SLAConfig.query.filter_by(test_name=self.test_name, is_active=True).first()
            if custom:
                return custom.sla_hours
        except Exception as e:
            print(f"Error getting SLA config: {e}")
        
        # Default fallback based on test name patterns
        if "STAT" in self.test_name or "Emergency" in self.test_name:
            return 1.0
        elif "Routine" in self.test_name:
            return 5.0
        elif "Microbiology" in self.test_name:
            return 72.0
        elif "Histology" in self.test_name:
            return 720.0
        elif "Special" in self.test_name:
            return 72.0
        else:
            return 24.0
    
    def update_sla_status(self):
        try:
            if self.result_ready_at:
                end_time = self.result_ready_at
            else:
                end_time = datetime.utcnow()
            
            tat = (end_time - self.created_at).total_seconds() / 3600
            self.tat_hours = round(tat, 2)
            
            sla_hours = self.get_sla_hours()
            
            if self.result_ready_at:
                if tat <= sla_hours:
                    self.sla_status = "ON TIME"
                elif tat <= sla_hours * 1.2:
                    self.sla_status = "NEAR DELAY"
                else:
                    self.sla_status = "DELAYED"
            else:
                if tat <= sla_hours:
                    self.sla_status = "IN PROGRESS"
                else:
                    self.sla_status = "DELAYED (ONGOING)"
            
            db.session.commit()
        except Exception as e:
            print(f"Error updating SLA status: {e}")


class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    
    id = db.Column(db.Integer, primary_key=True)
    case_id = db.Column(db.Integer, db.ForeignKey('cases.id'))
    action = db.Column(db.String(100))
    phase = db.Column(db.String(50))
    details = db.Column(db.Text)
    ip_address = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class PhaseComment(db.Model):
    __tablename__ = 'phase_comments'
    
    id = db.Column(db.Integer, primary_key=True)
    case_id = db.Column(db.Integer, db.ForeignKey('cases.id'))
    phase = db.Column(db.String(50))
    comment = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


# ========== HELPER FUNCTIONS ==========
def get_client_ip():
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0]
    return request.remote_addr

def log_audit(case_id, action, phase=None, details=None):
    try:
        audit = AuditLog(
            case_id=case_id,
            action=action,
            phase=phase,
            details=json.dumps(details) if details else None,
            ip_address=get_client_ip()
        )
        db.session.add(audit)
        db.session.commit()
    except Exception as e:
        print(f"Audit error: {e}")

def generate_case_id(dept):
    count = Case.query.filter_by(department=dept).count() + 1
    prefix = {"Lab": "LAB", "Radiology": "RAD", "Mobile": "MOB"}[dept]
    return f"{prefix}-{str(count).zfill(4)}"

def generate_visit_id():
    return f"VIS-{datetime.now().strftime('%Y%m%d%H%M%S')}"

def get_tests_by_department():
    """Get all active tests grouped by department from SLA Config"""
    configs = SLAConfig.query.filter_by(is_active=True).all()
    
    tests = {
        'Lab': [],
        'Radiology': [],
        'Mobile': []
    }
    
    for config in configs:
        if config.department in tests:
            tests[config.department].append(config.test_name)
    
    # Add default tests if none exist
    if not tests['Lab']:
        tests['Lab'] = [
            "STAT - Emergency",
            "Routine - Hematology Manual",
            "Routine - Serology",
            "Routine - Chemistry",
            "Specialized - Microbiology",
            "Specialized - Histology Cytology"
        ]
    
    if not tests['Radiology']:
        tests['Radiology'] = [
            "Radiology - X-ray Routine",
            "Radiology - X-ray Special",
            "Radiology - Ultrasound Routine",
            "Radiology - Ultrasound Special"
        ]
    
    if not tests['Mobile']:
        tests['Mobile'] = ["Mobile - Sample Collection & Transport"]
    
    return tests


# ========== SEED DEFAULT SLA CONFIGS ==========
def seed_default_sla_configs():
    """Add default SLA configurations if none exist"""
    defaults = [
        ("STAT - Emergency", "Lab", 1.0),
        ("Routine - Hematology Manual", "Lab", 5.0),
        ("Routine - Serology", "Lab", 5.0),
        ("Routine - Chemistry", "Lab", 5.0),
        ("Specialized - Microbiology", "Lab", 72.0),
        ("Specialized - Histology Cytology", "Lab", 720.0),
        ("Radiology - X-ray Routine", "Radiology", 5.0),
        ("Radiology - X-ray Special", "Radiology", 72.0),
        ("Radiology - Ultrasound Routine", "Radiology", 5.0),
        ("Radiology - Ultrasound Special", "Radiology", 24.0),
        ("Mobile - Sample Collection & Transport", "Mobile", 24.0),
    ]
    
    for test_name, dept, hours in defaults:
        existing = SLAConfig.query.filter_by(test_name=test_name).first()
        if not existing:
            config = SLAConfig(
                test_name=test_name,
                department=dept,
                sla_hours=hours,
                is_active=True
            )
            db.session.add(config)
    
    db.session.commit()


# ========== ROUTES ==========
@app.route('/')
def dashboard():
    department = request.args.get('dept')
    sla_filter = request.args.get('sla')
    search = request.args.get('search', '')
    
    query = Case.query
    
    if department:
        query = query.filter_by(department=department)
    if sla_filter:
        query = query.filter_by(sla_status=sla_filter)
    if search:
        query = query.filter(
            db.or_(
                Case.patient_name.contains(search),
                Case.case_id.contains(search)
            )
        )
    
    cases = query.order_by(Case.created_at.desc()).all()
    
    # Update SLA status for all cases
    for case in cases:
        case.update_sla_status()
    
    # Alerts
    alerts = [c for c in cases if c.sla_status == 'NEAR DELAY' or 
              (c.sla_status == 'IN PROGRESS' and c.tat_hours and 
               c.tat_hours > c.get_sla_hours() * 0.8)]
    
    # Stats
    stats = {
        'total': Case.query.count(),
        'delayed': Case.query.filter(Case.sla_status.like('%DELAYED%')).count(),
        'on_time': Case.query.filter_by(sla_status='ON TIME').count(),
        'in_progress': Case.query.filter(Case.status != 'Completed').count()
    }
    
    return render_template('dashboard.html', 
                         cases=cases, 
                         stats=stats,
                         alerts=alerts,
                         departments=['Lab', 'Radiology', 'Mobile'],
                         sla_statuses=['ON TIME', 'NEAR DELAY', 'DELAYED', 'IN PROGRESS', 'DELAYED (ONGOING)'],
                         current_dept=department,
                         current_sla=sla_filter,
                         search=search)

@app.route('/export/excel')
def export_excel():
    cases = Case.query.order_by(Case.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "TAT Cases"
    
    headers = ['Case ID', 'Patient Name', 'Department', 'Test Name', 'Status', 'SLA Status', 'TAT (hours)', 'Created At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for row, case in enumerate(cases, 2):
        ws.cell(row=row, column=1, value=case.case_id)
        ws.cell(row=row, column=2, value=case.patient_name)
        ws.cell(row=row, column=3, value=case.department)
        ws.cell(row=row, column=4, value=case.test_name)
        ws.cell(row=row, column=5, value=case.status)
        ws.cell(row=row, column=6, value=case.sla_status or 'PENDING')
        ws.cell(row=row, column=7, value=case.tat_hours or '')
        ws.cell(row=row, column=8, value=case.created_at.strftime('%Y-%m-%d %H:%M') if case.created_at else '')
    
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_audit(None, 'EXPORT_EXCEL', details={'count': len(cases)})
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'tat_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/new', methods=['GET', 'POST'])
def new_case():
    if request.method == 'POST':
        data = request.form
        
        case = Case(
            visit_id=generate_visit_id(),
            case_id=generate_case_id(data['department']),
            patient_name=data['patient_name'],
            department=data['department'],
            test_name=data['test_name'],
            test_category=data.get('test_category', 'Routine'),
            status='Not Started',
            sla_status='PENDING'
        )
        
        db.session.add(case)
        db.session.commit()
        
        log_audit(case.id, 'CREATE_CASE', details={'test_name': data['test_name']})
        
        return redirect(url_for('case_detail', case_id=case.id))
    
    # Get tests from SLA Config
    tests_by_dept = get_tests_by_department()
    
    return render_template('new.html', 
                         lab_tests=tests_by_dept['Lab'],
                         radiology_tests=tests_by_dept['Radiology'],
                         mobile_tests=tests_by_dept['Mobile'])

@app.route('/case/<int:case_id>')
def case_detail(case_id):
    case = Case.query.get_or_404(case_id)
    case.update_sla_status()
    
    department_phases = {
        "Lab": {
            "Pre-Analytical": ["sample_collection", "sample_labeling", "sample_transport", "sample_reception", "sample_processing"],
            "Analytical": ["analysis"],
            "Post-Analytical": ["review", "validation", "result_upload"]
        },
        "Radiology": {
            "Examination": ["exam"],
            "Post-Examination": ["post_exam"]
        },
        "Mobile": {
            "Registration": ["register"],
            "Collection": ["mobile_collect"],
            "Transport": ["transport_to_lab"],
            "Results": ["deliver_results"]
        }
    }
    
    phase_targets = {
        "sample_collection": 10, "sample_labeling": 3, "sample_transport": 30,
        "sample_reception": 6, "sample_processing": 5, "analysis": 90,
        "review": 10, "validation": 5, "result_upload": 4, "exam": 20,
        "post_exam": 2880, "register": 10, "mobile_collect": 10,
        "transport_to_lab": 60, "deliver_results": 30
    }
    
    phases = department_phases.get(case.department, {})
    comments = PhaseComment.query.filter_by(case_id=case_id).order_by(PhaseComment.created_at.desc()).all()
    audits = AuditLog.query.filter_by(case_id=case_id).order_by(AuditLog.created_at.desc()).limit(20).all()
    
    return render_template('case_detail.html', 
                         case=case, phases=phases, targets=phase_targets,
                         comments=comments, audits=audits)

@app.route('/api/start_phase/<int:case_id>/<phase>')
def start_phase(case_id, phase):
    case = Case.query.get_or_404(case_id)
    field_name = f"{phase}_start"
    
    if hasattr(case, field_name):
        setattr(case, field_name, datetime.utcnow())
        case.status = 'In Progress'
        db.session.commit()
        log_audit(case.id, 'START_PHASE', phase)
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Invalid phase'})

@app.route('/api/end_phase/<int:case_id>/<phase>')
def end_phase(case_id, phase):
    case = Case.query.get_or_404(case_id)
    field_name = f"{phase}_end"
    
    if hasattr(case, field_name):
        setattr(case, field_name, datetime.utcnow())
        
        # Check if all phases are complete
        dept_phases = {
            "Lab": ["sample_collection", "sample_labeling", "sample_transport", "sample_reception", "sample_processing", "analysis", "review", "validation", "result_upload"],
            "Radiology": ["exam", "post_exam"],
            "Mobile": ["register", "mobile_collect", "transport_to_lab", "deliver_results"]
        }
        
        remaining = []
        for p in dept_phases.get(case.department, []):
            if getattr(case, f"{p}_end") is None:
                remaining.append(p)
        
        if len(remaining) == 0:
            case.result_ready_at = datetime.utcnow()
            case.status = 'Completed'
        
        db.session.commit()
        case.update_sla_status()
        log_audit(case.id, 'END_PHASE', phase)
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Invalid phase'})

@app.route('/api/manual_duration/<int:case_id>', methods=['POST'])
def manual_duration(case_id):
    case = Case.query.get_or_404(case_id)
    data = request.json
    phase = data.get('phase')
    minutes = float(data.get('minutes'))
    
    case.set_manual_duration(phase, minutes)
    
    if not getattr(case, f"{phase}_start"):
        setattr(case, f"{phase}_start", datetime.utcnow())
        setattr(case, f"{phase}_end", datetime.utcnow())
        db.session.commit()
    
    log_audit(case.id, 'MANUAL_DURATION', phase, {'minutes': minutes})
    return jsonify({'success': True})

@app.route('/api/add_comment/<int:case_id>', methods=['POST'])
def add_comment(case_id):
    data = request.json
    comment = PhaseComment(
        case_id=case_id,
        phase=data.get('phase', 'general'),
        comment=data.get('comment')
    )
    db.session.add(comment)
    db.session.commit()
    log_audit(case_id, 'ADD_COMMENT', data.get('phase'), {'comment': data.get('comment')[:50]})
    return jsonify({'success': True})

@app.route('/bulk_action', methods=['POST'])
def bulk_action():
    data = request.json
    case_ids = data.get('case_ids', [])
    action = data.get('action')
    phase = data.get('phase')
    
    success_count = 0
    for case_id in case_ids:
        case = Case.query.get(case_id)
        if case:
            if action == 'start_phase':
                field_name = f"{phase}_start"
                if hasattr(case, field_name) and not getattr(case, field_name):
                    setattr(case, field_name, datetime.utcnow())
                    case.status = 'In Progress'
                    success_count += 1
            elif action == 'complete_phase':
                field_name = f"{phase}_end"
                if hasattr(case, field_name) and not getattr(case, field_name):
                    setattr(case, field_name, datetime.utcnow())
                    success_count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'updated': success_count})

@app.route('/audit_log')
def audit_log_page():
    page = request.args.get('page', 1, type=int)
    per_page = 50
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).paginate(page=page, per_page=per_page)
    return render_template('audit_log.html', logs=logs)

@app.route('/sla_config')
def sla_config_page():
    configs = SLAConfig.query.order_by(SLAConfig.department, SLAConfig.test_name).all()
    departments = ['Lab', 'Radiology', 'Mobile']
    return render_template('sla_config.html', configs=configs, departments=departments)

@app.route('/api/sla_config', methods=['POST'])
def update_sla_config():
    data = request.json
    test_name = data.get('test_name')
    department = data.get('department')
    sla_hours = float(data.get('sla_hours'))
    
    config = SLAConfig.query.filter_by(test_name=test_name).first()
    if config:
        config.sla_hours = sla_hours
        config.department = department
        config.updated_at = datetime.utcnow()
    else:
        config = SLAConfig(
            test_name=test_name,
            department=department,
            sla_hours=sla_hours,
            is_active=True
        )
        db.session.add(config)
    
    db.session.commit()
    log_audit(None, 'UPDATE_SLA_CONFIG', details={'test_name': test_name, 'department': department, 'sla_hours': sla_hours})
    
    return jsonify({'success': True})

@app.route('/api/sla_config/<int:config_id>', methods=['DELETE'])
def delete_sla_config(config_id):
    config = SLAConfig.query.get_or_404(config_id)
    db.session.delete(config)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/sla_config/<int:config_id>/toggle', methods=['POST'])
def toggle_sla_config(config_id):
    config = SLAConfig.query.get_or_404(config_id)
    config.is_active = not config.is_active
    config.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True, 'is_active': config.is_active})


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        seed_default_sla_configs()  # Add default SLA configs
    print("✅ Database initialized with default SLA configurations!")
    print("👉 Open http://localhost:5000")
    app.run(host='0.0.0.0', port=5000, debug=True)